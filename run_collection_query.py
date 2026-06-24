import argparse
import csv
import importlib
import os
import re
import subprocess
import sys
from datetime import datetime
from decimal import Decimal
from pathlib import Path


def ensure_python_package(import_name, package_name=None):
    try:
        return importlib.import_module(import_name)
    except ImportError:
        package_name = package_name or import_name
        print(f"STATUS: Missing Python package '{package_name}'. Installing now...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])
        except subprocess.CalledProcessError as exc:
            print(f"STATUS: Failed to install Python package '{package_name}'.")
            raise exc
        print(f"STATUS: Python package '{package_name}' installed successfully.")
        return importlib.import_module(import_name)


fdb = ensure_python_package("fdb")
ensure_python_package("openpyxl")
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DEFAULT_DB_PATHS = [
    r"main-server:i_tax046zamboanguita",
    # r"E:\ZAMBOANGUITA.FDB",
    # r"C:\ZAMBOANGUITA_DB\ZAMBOANGUITA.FDB",
]
DEFAULT_FB_CLIENT_PATHS = [
    r"C:\Program Files\Firebird\Firebird_2_5\bin\fbclient.dll",
    r"C:\ITAX\utilities\fbclient.dll",
]
DEFAULT_ODBC_DSN = "itaxzamboanguita"
CONNECTION_MODE = os.environ.get("ESRE_CONNECTION", "odbc").strip().lower()
SQL_FILE = Path(__file__).resolve().parent / "firebird_metadata" / "collection_analysis_queries.sql"
OUTPUT_DIR = Path(__file__).resolve().parent / "firebird_metadata" / "output"
GOOGLE_EXPORT_DIR = Path(__file__).resolve().parent / "google_sheet_exports"
TEMPLATE_DIR = Path(__file__).resolve().parent / "report_template"
BUSINESS_PERMIT_DIR = Path(__file__).resolve().parent / "BUSINESS_PERMIT_REPORT"


def is_firebird_server_path(value):
    value = (value or "").strip()
    if not value:
        return False
    if re.match(r"^[A-Za-z]:[\\/]", value):
        return False
    return ":" in value


def db_path_candidates():
    env_path = os.environ.get("ESRE_FIREBIRD_DB")
    if env_path:
        return [env_path]
    return list(DEFAULT_DB_PATHS)


def resolve_db_path():
    candidates = db_path_candidates()
    for candidate in candidates:
        candidate = (candidate or "").strip()
        if not candidate:
            continue
        if is_firebird_server_path(candidate):
            return candidate
        if Path(candidate).exists():
            return candidate
    return candidates[0] if candidates else DEFAULT_DB_PATHS[0]


def resolve_fb_client():
    env_path = os.environ.get("ESRE_FIREBIRD_CLIENT")
    candidates = [env_path] if env_path else []
    candidates.extend(DEFAULT_FB_CLIENT_PATHS)
    for candidate in candidates:
        candidate = (candidate or "").strip()
        if candidate and Path(candidate).exists():
            return candidate
    return candidates[0] if candidates else DEFAULT_FB_CLIENT_PATHS[0]


def resolve_odbc_dsn():
    return os.environ.get("ESRE_ODBC_DSN", DEFAULT_ODBC_DSN).strip()
FDB_TEMPLATE_REPORTS = {
    21: "Summary of Collection",
    22: "Summary of Collection no rpt",
    23: "Summary of Collection rpt",
    24: "Summary in RPT based on the SUMMARY sheet layout",
    25: "Record of Real Property Tax Collection",
    26: "Record of Real Property Tax Collection - Advance Payment Report",
    27: "Summary Report Sharing",
    28: "Provincial RPT Coding / Province Remittance Report",
    29: "Abstract of General Collections",
    30: "Abstract of Trust Funds Collections",
    31: "Full Report Collections",
    32: "CMCI Annex A-B Business Permit Registration Report",
    33: "Tax on Business Summary from BPLS Business Tax",
    34: "Generate Receipt Collector",
}

COLLECTOR_ALIASES = {
    "iris": "angelique",
    "iris arbolado": "angelique",
    "angelique iris": "angelique",
    "flora my": "flora",
}


def parse_date(value):
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", value):
        raise argparse.ArgumentTypeError("Date must use YYYY-MM-DD format.")
    return value


def load_queries():
    text = SQL_FILE.read_text(encoding="utf-8")
    matches = list(re.finditer(r"/\*\s*(\d+)\.\s+(.+?)\s*\*/", text, flags=re.S))
    queries = []

    for index, match in enumerate(matches):
        number = int(match.group(1))
        title = " ".join(match.group(2).split())
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        sql = text[start:end].strip()
        if sql.endswith(";"):
            sql = sql[:-1].strip()
        queries.append({"number": number, "title": title, "sql": sql})

    return queries


def list_queries(queries):
    for query in queries:
        print(f"{query['number']:>2}. {query['title']}")
    for number, title in FDB_TEMPLATE_REPORTS.items():
        print(f"{number:>2}. {title}")


def sql_literal(value):
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def apply_parameters(sql, date_from, date_to, collector=None):
    return (
        sql.replace(":date_from", sql_literal(date_from))
        .replace(":date_to", sql_literal(date_to))
        .replace(":collector", sql_literal(collector))
    )


def export_rows(cursor, output_path):
    columns = [description[0].strip() for description in cursor.description]
    count = 0

    try:
        handle = output_path.open("w", newline="", encoding="utf-8-sig")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        handle = output_path.open("w", newline="", encoding="utf-8-sig")

    with handle:
        writer = csv.writer(handle)
        writer.writerow(columns)
        while True:
            rows = cursor.fetchmany(1000)
            if not rows:
                break
            writer.writerows(rows)
            count += len(rows)

    return count, output_path


def cell_value(cell):
    value = cell.value
    if value is None:
        return ""
    return value


def write_csv_rows(output_path, rows):
    try:
        handle = output_path.open("w", newline="", encoding="utf-8-sig")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        handle = output_path.open("w", newline="", encoding="utf-8-sig")

    with handle:
        writer = csv.writer(handle)
        writer.writerows(rows)
    return len(rows) - 1 if rows else 0, output_path


def excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def save_workbook_with_fallback(workbook, output_path):
    try:
        workbook.save(output_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        workbook.save(output_path)
    return output_path


def write_template_workbook(number, rows, output_path, date_from):
    template_map = {
        21: ("summary_of_collection_template.xlsx", 8),
        22: ("summary_of_collection_template_no_rpt.xlsx", 8),
        23: ("summary_of_collection_template_rpt.xlsx", 8),
    }
    template_name, start_row = template_map[number]
    workbook = load_workbook(TEMPLATE_DIR / template_name)
    sheet = workbook.active

    try:
        month_label = datetime.strptime(date_from, "%Y-%m-%d").strftime("Month of %B %Y")
        sheet["A4"] = month_label
    except ValueError:
        pass

    body_rows = rows[1:]
    for index, row_values in enumerate(body_rows):
        excel_row = start_row + index
        for col_index, value in enumerate(row_values[:12], start=1):
            sheet.cell(excel_row, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(body_rows), output_path


def write_rpt_record_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "RECORD OF REAL PROPERTY TAX COLLECTION.xlsx")
    sheet = workbook.active

    try:
        start_label = datetime.strptime(date_from, "%Y-%m-%d").strftime("%B %d, %Y")
        end_label = datetime.strptime(date_to, "%Y-%m-%d").strftime("%B %d, %Y")
        sheet["E4"] = start_label if date_from == date_to else f"{start_label} to {end_label}"
    except ValueError:
        sheet["E4"] = f"{date_from} to {date_to}"

    start_row = 11
    body_rows = rows[1:]
    for index, row_values in enumerate(body_rows):
        excel_row = start_row + index
        for col_index, value in enumerate(row_values[:36], start=1):
            sheet.cell(excel_row, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(body_rows), output_path


def write_advance_rpt_record_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(
        TEMPLATE_DIR / "RECORD OF REAL PROPERTY TAX COLLECTION - ADVANCE PAYMENT REPORT.xlsx"
    )
    sheet = workbook.active

    try:
        start_label = datetime.strptime(date_from, "%Y-%m-%d").strftime("%B %d, %Y")
        end_label = datetime.strptime(date_to, "%Y-%m-%d").strftime("%B %d, %Y")
        sheet["E4"] = start_label if date_from == date_to else f"{start_label} to {end_label}"
    except ValueError:
        sheet["E4"] = f"{date_from} to {date_to}"

    start_row = 11
    body_rows = rows[1:]
    for index, row_values in enumerate(body_rows):
        excel_row = start_row + index
        for col_index, value in enumerate(row_values[:18], start=1):
            sheet.cell(excel_row, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(body_rows), output_path


def write_summary_sharing_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "SUMMARY_REPORT_SHARING_TEMPLATE.xlsx")
    sheet = workbook.active

    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        sheet["F3"] = start_date.strftime("%B") if start_date.month == end_date.month else (
            f"{start_date.strftime('%B')} to {end_date.strftime('%B')}"
        )
        sheet["F4"] = str(start_date.day) if date_from == date_to else f"{start_date.day}-{end_date.day}"
        sheet["F5"] = str(end_date.year)
    except ValueError:
        sheet["F3"] = date_from
        sheet["F4"] = date_to

    for row_index, col_index, value in rows[1:]:
        sheet.cell(row_index, col_index).value = excel_value(value)

    formula_updates = {
        "J36": "=J15-K15",
        "K36": "=(J15-K15)*0.5",
        "L36": "=(J15-K15)*0.5",
        "J37": "=L15",
        "K37": "=L15*0.5",
        "L37": "=L15*0.5",
        "J38": "=SUM(M15:N15)",
        "K38": "=SUM(M15:N15)*0.5",
        "L38": "=SUM(M15:N15)*0.5",
        "J46": "=J27-K27",
        "K46": "=(J27-K27)*0.5",
        "L46": "=(J27-K27)*0.5",
        "J47": "=L27",
        "K47": "=L27*0.5",
        "L47": "=L27*0.5",
        "J48": "=SUM(M27:N27)",
        "K48": "=SUM(M27:N27)*0.5",
        "L48": "=SUM(M27:N27)*0.5",
        "C51": "=SUM(D49:F49)",
    }
    for cell, formula in formula_updates.items():
        sheet[cell] = formula

    sheet["C45"] = "BUILDING"
    sheet["J45"] = "BUILDING"
    sheet["B26"] = "BLDG-INDUS/SPECIAL"
    sheet["I26"] = "BLDG-INDUS/SPECIAL"
    sheet["I51"] = "BUILDING SHARING TOTAL"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(rows) - 1 if rows else 0, output_path


def period_label(date_from, date_to):
    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        start_label = start_date.strftime("%B %d, %Y")
        end_label = end_date.strftime("%B %d, %Y")
        return start_label if date_from == date_to else f"{start_label} to {end_label}"
    except ValueError:
        return f"{date_from} to {date_to}"


def write_provincial_rpt_coding_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "PROVINCIAL_RPT_CODING_TEMPLATE.xlsx")

    for sheet_name in ("GF", "SEF"):
        sheet = workbook[sheet_name]
        sheet["E5"] = period_label(date_from, date_to)
        for row_index in range(9, 22):
            for col_index in (3, 5, 7, 9):
                sheet.cell(row_index, col_index).value = 0

    for sheet_name, row_index, col_index, value in rows[1:]:
        workbook[sheet_name].cell(row_index, col_index).value = excel_value(value)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(rows) - 1 if rows else 0, output_path


def write_abstract_general_collections_workbook(rows, daily_rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "ABSTRACT_OF_GENERAL_COLLECTIONS.xlsx")
    data_sheet = workbook["data"]
    daily_sheet = workbook["daily_collection"]
    data_sheet["K4"] = period_label(date_from, date_to)

    for index, row_values in enumerate(rows[1:], start=8):
        for col_index, value in enumerate(row_values, start=1):
            data_sheet.cell(index, col_index).value = excel_value(value)

    for index, row_values in enumerate(daily_rows[1:], start=5):
        for col_index, value in enumerate(row_values, start=1):
            daily_sheet.cell(index, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(rows) - 1 if rows else 0, output_path


def write_abstract_trust_funds_workbook(rows, daily_rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "ABSTRACT_OF_TRUST_FUNDS_COLLECTIONS.xlsx")
    data_sheet = workbook["data"]
    daily_sheet = workbook["daily_collection"]
    data_sheet["H4"] = period_label(date_from, date_to)

    for index, row_values in enumerate(rows[1:], start=9):
        for col_index, value in enumerate(row_values, start=1):
            data_sheet.cell(index, col_index).value = excel_value(value)

    for index, row_values in enumerate(daily_rows[1:], start=5):
        for col_index, value in enumerate(row_values, start=1):
            daily_sheet.cell(index, col_index).value = excel_value(value)

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(rows) - 1 if rows else 0, output_path


def write_full_report_collections_workbook(rows, output_path, date_from, date_to):
    workbook = load_workbook(TEMPLATE_DIR / "FULL_REPORT_COLLECTIONS.xlsx")
    sheet = workbook.active

    try:
        start_date = datetime.strptime(date_from, "%Y-%m-%d")
        end_date = datetime.strptime(date_to, "%Y-%m-%d")
        sheet["D4"] = start_date.strftime("%B") if start_date.month == end_date.month else (
            f"{start_date.strftime('%B')} to {end_date.strftime('%B')}"
        )
        sheet["D5"] = str(end_date.year)
    except ValueError:
        sheet["D4"] = date_from
        sheet["D5"] = date_to

    body_rows = rows[1:]
    start_row = 8
    template_last_daily_row = 31
    capacity = template_last_daily_row - start_row + 1
    if len(body_rows) > capacity:
        sheet.insert_rows(template_last_daily_row + 1, len(body_rows) - capacity)

    total_row = start_row + max(len(body_rows), capacity)
    summary_row_1 = total_row + 4
    summary_row_2 = total_row + 5
    summary_row_3 = total_row + 6

    for excel_row in range(start_row, total_row):
        for col_index in range(1, 7):
            sheet.cell(excel_row, col_index).value = None

    for index, row_values in enumerate(body_rows):
        excel_row = start_row + index
        for col_index, value in enumerate(row_values[:4], start=1):
            sheet.cell(excel_row, col_index).value = excel_value(value)
        sheet.cell(excel_row, 5).value = None
        sheet.cell(excel_row, 6).value = f"=SUM(B{excel_row}:D{excel_row})"

    sheet.cell(total_row, 1).value = "TOTAL"
    for col_letter in ("B", "C", "D", "E", "F"):
        sheet[f"{col_letter}{total_row}"] = f"=SUM({col_letter}{start_row}:{col_letter}{total_row - 1})"

    sheet[f"C{summary_row_1}"] = "RCD TOTAL"
    sheet[f"F{summary_row_1}"] = f"=F{total_row}"
    sheet[f"C{summary_row_2}"] = "LESS: DUE FROM"
    sheet[f"F{summary_row_2}"] = f"=E{total_row}"
    sheet[f"C{summary_row_3}"] = "TOTAL COLLECTIONS"
    sheet[f"F{summary_row_3}"] = f"=F{summary_row_1}-F{summary_row_2}"

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(body_rows), output_path


def build_cmci_annex_mapping_rows(date_from, date_to):
    return [
        ["SECTION", "FIELD", "SOURCE / RULE", "STATUS"],
        ["Report", "Report No.", "32", "Added to list"],
        [
            "Report",
            "Report Name",
            "CMCI Annex A-B Business Permit Registration Report",
            "Analyzed / pending final data mapping",
        ],
        [
            "Template",
            "Workbook",
            "BUSINESS_PERMIT_REPORT\\2025-2026_ANNEX-A-B_cmci_report.xlsx",
            "Analyzed only",
        ],
        [
            "Template",
            "Sheets",
            "Annex A (Jan. to Dec. 2025), Annex B (Jan. to Mar. 2026), PSIC",
            "Observed",
        ],
        ["Template", "Period Requested", f"{date_from} to {date_to}", "Input parameter"],
        ["Column A", "LGU", "Municipality of Zamboanguita", "Static / confirm exact CMCI spelling"],
        ["Column B", "Province", "Negros Oriental", "Static"],
        ["Column C", "Region", "REGION VII (CENTRAL VISAYAS)", "Static"],
        ["Column D", "Classification", "Third Class Municipality", "Static"],
        ["Column E", "LGU Type", "Municipality", "Manual/static value needed"],
        ["Column F", "Business Name", "REGISTERED_BUSINESSES-BPLS / BUSINESS_ESTABLISHMENT-BPLS", "Source mapped"],
        [
            "Columns G-I",
            "Business Address",
            "Split address into house/building, street/barangay, subdivision/district where available",
            "Needs parsing rule",
        ],
        ["Column J", "Owner's Name", "Owner/applicant name fields from BPLS", "Source mapped"],
        [
            "Column K",
            "Industry / Nature of Business",
            "Map BPLS line of business or business line code to PSIC major category",
            "Needs PSIC mapping table",
        ],
        [
            "Column L",
            "Business Type",
            "Normalize SOLE PROPRIETORSHIP, CORPORATION, PARTNERSHIP, COOPERATIVE, OPC",
            "Needs value normalization",
        ],
        [
            "Column M",
            "Capitalization Size",
            "Compute CMCI dropdown value with threshold text from capitalization or gross fallback",
            "Source mapped",
        ],
        ["Column N", "New / Renewal", "Map registration status NEW or RENEWAL", "Source mapped"],
        ["Column O", "Year of Registration", "Use registration/application year for Annex A/B period", "Source mapped"],
        ["Column P", "Permit No.", "Permit number from BPLS establishment/registered data", "Source mapped"],
        [
            "Next Step",
            "PSIC mapping",
            "Create reviewed mapping from BPLS business lines to the 21 PSIC major categories",
            "Required before real Excel generation",
        ],
    ]


def find_business_permit_workbook(pattern):
    matches = sorted(BUSINESS_PERMIT_DIR.glob(pattern), key=lambda path: path.stat().st_mtime, reverse=True)
    if not matches:
        raise FileNotFoundError(f"Business permit source workbook not found: {BUSINESS_PERMIT_DIR / pattern}")
    return matches[0]


def parse_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y"):
            try:
                return datetime.strptime(value[:10] if fmt == "%Y-%m-%d" else value, fmt).date()
            except ValueError:
                pass
    return None


def load_sheet_records(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        records.append({headers[index]: value for index, value in enumerate(row) if index < len(headers)})
    workbook.close()
    return records


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def normalize_business_type(value):
    text = clean_text(value).upper()
    if "ONE" in text and "CORPORATION" in text:
        return "One-Person Corporation"
    if "CORPORATION" in text:
        return "Corporation"
    if "PARTNERSHIP" in text:
        return "Partnership"
    if "COOPERATIVE" in text:
        return "Cooperative"
    return "Single Proprietor"


def capitalization_size(capital, gross):
    basis = Decimal("0")
    for value in (capital, gross):
        if value in (None, ""):
            continue
        try:
            basis = Decimal(str(value))
        except Exception:
            basis = Decimal("0")
        if basis > 0:
            break

    if basis <= Decimal("3000000"):
        return "Micro (less than P3000000)"
    if basis <= Decimal("15000000"):
        return "Small ( P3000001 - P15000000)"
    if basis <= Decimal("100000000"):
        return "Medium (P15000001 - P100000000)"
    return "Large (more than P100000000)"


def normalize_new_renewal(value):
    text = clean_text(value).upper()
    return "New" if text == "NEW" else "Renewal"


def psic_category(line_of_business):
    text = clean_text(line_of_business).lower()
    rules = [
        ("Financial and Insurance Activities", ("bank", "financial", "lending", "pawn", "money", "remittance", "insurance")),
        ("Real Estate Activities", ("real estate", "lessor", "apartment", "rental", "space for rent", "property")),
        (
            "Wholesale and Retail Trade; Repair of Motor Vehicles and Motorcycles",
            ("retail", "wholesale", "store", "sari-sari", "sale of", "pharmacy", "hardware", "lpg", "gasoline", "motorcycle parts"),
        ),
        (
            "Accommodation and Food Service Activities",
            ("restaurant", "eatery", "carinderia", "food", "cafe", "coffee", "hotel", "resort", "accommodation", "guesthouse", "lodging", "catering"),
        ),
        ("Transportation and Storage", ("transport", "tricycle", "pedicab", "vehicle", "passenger", "cargo", "trucking")),
        ("Manufacturing", ("manufactur", "baking", "bakery", "milling", "printing", "processed")),
        ("Agriculture, Forestry And Fishing", ("agri", "farm", "crop", "forestry", "coconut", "livestock", "poultry", "fishing")),
        ("Mining and Quarrying", ("mining", "quarry", "sand", "gravel")),
        (
            "Water Supply; Sewerage, Waste Management And Remediation Activities",
            ("water", "refilling", "purifying", "waste", "sewerage"),
        ),
        ("Construction", ("construction", "contractor", "building")),
        ("Information and Communication", ("internet", "pisonet", "computer", "telecom", "communication")),
        ("Professional, Scientific and Technical Activities", ("legal", "accounting", "engineering", "consult", "technical")),
        ("Administrative and Support Service Activities", ("travel agency", "security", "manpower", "support service")),
        ("Education", ("school", "tutorial", "education", "training")),
        ("Human Health and Social Work Activities", ("clinic", "medical", "dental", "laboratory", "health")),
        ("Arts, Entertainment and Recreation", ("gambling", "cockpit", "amusement", "recreation", "sports")),
        ("Other Service Activities", ("barber", "beauty", "salon", "funeral", "wellness", "repair", "personal service")),
    ]
    for category, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return category
    return "Other Service Activities"


def establishment_lookup(records, date_from, date_to):
    start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
    end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    by_business_id = {}

    for record in records:
        business_id = clean_text(record.get("Business Identification Number"))
        if not business_id:
            continue
        application_date = parse_excel_date(record.get("Application Date"))
        in_range = application_date is not None and start_date <= application_date <= end_date
        score = (
            1 if in_range else 0,
            1 if clean_text(record.get("Permit No.")) else 0,
            1 if (record.get("Total Amount Paid") or 0) else 0,
        )
        existing_score, _existing_record = by_business_id.get(business_id, ((-1, -1, -1), None))
        if score > existing_score:
            by_business_id[business_id] = (score, record)

    return {business_id: record for business_id, (_score, record) in by_business_id.items()}


def cmci_annex_row(registered_record, establishment_record):
    capital = registered_record.get("Capital Investment")
    gross = (registered_record.get("Gross Essential") or 0) + (registered_record.get("Gross Non-essential") or 0)
    if establishment_record:
        capital = capital or establishment_record.get("Capital Investment")
        gross = gross or establishment_record.get("Gross Sales")

    line_of_business = clean_text(
        registered_record.get("Line of Business")
        or (establishment_record or {}).get("Business Line")
        or (establishment_record or {}).get("Business Nature")
    )
    address = clean_text(registered_record.get("Business Address") or (establishment_record or {}).get("Location of Business"))
    barangay = clean_text(registered_record.get("Barangay Name") or (establishment_record or {}).get("Barangay (Business Address)"))
    permit_no = clean_text(registered_record.get("Business Permit No.") or (establishment_record or {}).get("Permit No."))
    date_applied = parse_excel_date(registered_record.get("Date Applied") or (establishment_record or {}).get("Application Date"))
    year = registered_record.get("Year") or (date_applied.year if date_applied else "")

    return [
        "Zamboanguita",
        "Negros Oriental",
        "REGION VII (CENTRAL VISAYAS)",
        "Third Class Municipality",
        "Municipality",
        clean_text(registered_record.get("Business Name") or (establishment_record or {}).get("Business Name")),
        "",
        address or barangay,
        "",
        clean_text(registered_record.get("Name of Owner/Applicant")),
        psic_category(line_of_business),
        normalize_business_type(registered_record.get("Type of Business") or (establishment_record or {}).get("Type of Business")),
        capitalization_size(capital, gross),
        normalize_new_renewal(registered_record.get("Status of Registration") or (establishment_record or {}).get("Type of Application")),
        year,
        permit_no,
    ]


def cmci_target_sheet_name(workbook, year):
    if year == 2025:
        return "Annex A (Jan. to Dec. 2025)"
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("Annex B"):
            return sheet_name
    return workbook.sheetnames[1]


def build_cmci_annex_rows(date_from, date_to):
    registered_path = find_business_permit_workbook("REGISTERED_BUSINESSES-BPLS*.xlsx")
    establishment_path = find_business_permit_workbook("BUSINESS_ESTABLISHMENT-BPLS*.xlsx")
    registered_records = load_sheet_records(registered_path)
    establishment_records = load_sheet_records(establishment_path)
    establishments = establishment_lookup(establishment_records, date_from, date_to)
    start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
    end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
    rows_by_year = {}

    for record in registered_records:
        date_applied = parse_excel_date(record.get("Date Applied"))
        if date_applied is None or not (start_date <= date_applied <= end_date):
            continue
        permit_no = clean_text(record.get("Business Permit No."))
        if not permit_no:
            continue
        address = clean_text(record.get("Business Address"))
        if address and "ZAMBOANGUITA" not in address.upper():
            continue
        business_id = clean_text(record.get("Business Identification Number"))
        row = cmci_annex_row(record, establishments.get(business_id))
        rows_by_year.setdefault(date_applied.year, []).append((date_applied, permit_no, row))

    for year in rows_by_year:
        rows_by_year[year].sort(key=lambda item: (item[0], item[1], item[2][5]))
    return rows_by_year


def write_cmci_annex_workbook(rows_by_year, output_path, date_from, date_to):
    template_path = BUSINESS_PERMIT_DIR / "2025-2026_ANNEX-A-B_cmci_report.xlsx"
    workbook = load_workbook(template_path)
    start_row = 7
    end_clear_row = 1301

    for year, rows in rows_by_year.items():
        sheet = workbook[cmci_target_sheet_name(workbook, year)]
        for row_index in range(start_row, min(sheet.max_row, end_clear_row) + 1):
            for col_index in range(1, 17):
                sheet.cell(row_index, col_index).value = None
        if len(rows) > sheet.max_row - start_row + 1:
            sheet.insert_rows(sheet.max_row + 1, len(rows) - (sheet.max_row - start_row + 1))
        for index, (_date_applied, _permit_no, row_values) in enumerate(rows, start=start_row):
            for col_index, value in enumerate(row_values, start=1):
                sheet.cell(index, col_index).value = excel_value(value)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return sum(len(rows) for rows in rows_by_year.values()), output_path


def load_sheet_records_with_header(path, header_row):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=header_row, max_row=header_row))]
    records = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        records.append(record)
    workbook.close()
    return records


def decimal_value(value):
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")


def tax_on_business_category(business_nature, business_line):
    nature_text = clean_text(business_nature).lower()
    line_text = clean_text(business_line).lower()
    text = f"{nature_text} {line_text}"
    if any(keyword in text for keyword in ("bank", "financial", "lending", "pawn", "money", "remittance", "insurance")):
        return "Banks & Other Financial Int."
    if any(keyword in text for keyword in ("manufactur", "baking", "bakery", "milling", "printing", "processed")):
        return "Manufacturing"
    if any(keyword in line_text for keyword in ("wholesale", "distributor", "distribution")):
        return "Distributor"
    if any(keyword in line_text for keyword in ("retail", "store", "sari-sari", "pharmacy", "hardware", "convenience")):
        return "Retailing"
    if "wholesale and retail trade" in nature_text:
        return "Retailing"
    return "Other Business Tax"


def business_establishment_match_lookup(records):
    by_or = {}
    by_business_id = {}

    for record in records:
        or_number = clean_text(record.get("OR Number"))
        business_id = clean_text(record.get("Business Identification Number"))
        has_paid = decimal_value(record.get("Total Amount Paid")) > 0
        has_permit = bool(clean_text(record.get("Permit No.")))
        score = (1 if has_paid else 0, 1 if has_permit else 0)

        if or_number:
            existing_score, _existing_record = by_or.get(or_number, ((-1, -1), None))
            if score > existing_score:
                by_or[or_number] = (score, record)

        if business_id:
            existing_score, _existing_record = by_business_id.get(business_id, ((-1, -1), None))
            if score > existing_score:
                by_business_id[business_id] = (score, record)

    return (
        {or_number: record for or_number, (_score, record) in by_or.items()},
        {business_id: record for business_id, (_score, record) in by_business_id.items()},
    )


def build_tax_on_business_report(date_from, date_to):
    abstract_path = find_business_permit_workbook("ABSTRACT_OF_GENERAL_COLLECTION-BPLS*.xlsx")
    establishment_path = find_business_permit_workbook("BUSINESS_ESTABLISHMENT-BPLS*.xlsx")
    abstract_records = load_sheet_records_with_header(abstract_path, 7)
    establishment_records = load_sheet_records(establishment_path)
    by_or, by_business_id = business_establishment_match_lookup(establishment_records)
    start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
    end_date = datetime.strptime(date_to, "%Y-%m-%d").date()

    category_order = [
        "Manufacturing",
        "Distributor",
        "Retailing",
        "Banks & Other Financial Int.",
        "Other Business Tax",
        "Fines & Penalties",
    ]
    summary = {
        category: {
            "business_tax": Decimal("0"),
            "surcharge": Decimal("0"),
            "receipt_numbers": set(),
            "business_ids": set(),
        }
        for category in category_order
    }
    details = []

    for record in abstract_records:
        or_date = parse_excel_date(record.get("O.R. Date"))
        if or_date is None or not (start_date <= or_date <= end_date):
            continue

        business_tax = decimal_value(record.get("Business Tax"))
        surcharge = decimal_value(record.get("Surcharge"))
        if business_tax == 0 and surcharge == 0:
            continue

        or_number = clean_text(record.get("O.R. Number"))
        business_id = clean_text(record.get("Business Identification Number"))
        establishment = by_or.get(or_number) or by_business_id.get(business_id) or {}
        business_nature = clean_text(establishment.get("Business Nature"))
        business_line = clean_text(establishment.get("Business Line"))
        category = tax_on_business_category(business_nature, business_line)
        match_basis = "OR Number" if or_number in by_or else ("Business ID" if business_id in by_business_id else "Unmatched")

        if business_tax:
            summary[category]["business_tax"] += business_tax
            summary[category]["receipt_numbers"].add(or_number)
            summary[category]["business_ids"].add(business_id)

        if surcharge:
            summary["Fines & Penalties"]["surcharge"] += surcharge
            summary["Fines & Penalties"]["receipt_numbers"].add(or_number)
            summary["Fines & Penalties"]["business_ids"].add(business_id)

        details.append([
            or_date,
            clean_text(record.get("Date Paid")),
            or_number,
            clean_text(record.get("Transaction Type")),
            business_id,
            clean_text(record.get("Business Name")),
            clean_text(record.get("Barangay Name")),
            business_nature,
            business_line,
            category,
            business_tax,
            surcharge,
            decimal_value(record.get("Amount Paid")),
            match_basis,
        ])

    summary_rows = []
    for category in category_order:
        business_tax = summary[category]["business_tax"]
        surcharge = summary[category]["surcharge"]
        summary_rows.append([
            category,
            business_tax,
            surcharge,
            business_tax + surcharge,
        ])

    return {
        "summary": summary_rows,
        "details": details,
        "abstract_path": str(abstract_path),
        "establishment_path": str(establishment_path),
    }


def tax_on_business_summary_amounts(date_from, date_to):
    report_data = build_tax_on_business_report(date_from, date_to)
    return {
        row[0]: row[3]
        for row in report_data["summary"]
    }


def style_tax_on_business_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == "TOTAL":
            for cell in row:
                cell.fill = total_fill
                cell.font = Font(bold=True)


def write_tax_on_business_workbook(report_data, output_path, date_from, date_to):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    detail_sheet = workbook.create_sheet("Detail")
    notes_sheet = workbook.create_sheet("Notes")

    summary_headers = ["Category", "Business Tax", "Fines & Penalties / Surcharge", "Total"]
    summary_sheet.append(summary_headers)
    for row in report_data["summary"]:
        summary_sheet.append([excel_value(value) for value in row])
    total_row = summary_sheet.max_row + 1
    summary_sheet.append([
        "TOTAL",
        f"=SUM(B2:B{total_row - 1})",
        f"=SUM(C2:C{total_row - 1})",
        f"=SUM(D2:D{total_row - 1})",
    ])

    detail_headers = [
        "O.R. Date",
        "Date Paid",
        "O.R. Number",
        "Transaction Type",
        "Business ID",
        "Business Name",
        "Barangay",
        "Business Nature",
        "Business Line",
        "Tax Category",
        "Business Tax",
        "Surcharge",
        "Amount Paid",
        "Match Basis",
    ]
    detail_sheet.append(detail_headers)
    for row in report_data["details"]:
        detail_sheet.append([excel_value(value) for value in row])

    notes = [
        ["Report", "33. Tax on Business Summary from BPLS Business Tax"],
        ["Period", f"{date_from} to {date_to}"],
        ["General Collection Source", report_data["abstract_path"]],
        ["Business Establishment Source", report_data["establishment_path"]],
        ["Business Tax", "Uses Abstract of General Collection column: Business Tax"],
        ["Fines & Penalties", "Uses Abstract of General Collection column: Surcharge"],
        ["Classification Source", "Business Establishment columns: Business Nature and Business Line"],
        ["Manual Review", "Manufacturing, Distributor, Retailing, Banks/Financial, and Other Business Tax can still be manually reviewed against the other app/database."],
    ]
    for row in notes:
        notes_sheet.append(row)

    detail_sheet.sheet_state = "hidden"
    notes_sheet.sheet_state = "hidden"

    for sheet in (summary_sheet, detail_sheet, notes_sheet):
        style_tax_on_business_sheet(sheet)
        sheet.freeze_panes = "A2"

    for column, width in {
        "A": 28, "B": 18, "C": 30, "D": 18,
    }.items():
        summary_sheet.column_dimensions[column].width = width

    detail_widths = {
        "A": 13, "B": 20, "C": 20, "D": 14, "E": 24, "F": 38, "G": 18,
        "H": 44, "I": 44, "J": 28, "K": 14, "L": 14, "M": 14, "N": 14,
    }
    for column, width in detail_widths.items():
        detail_sheet.column_dimensions[column].width = width
    notes_sheet.column_dimensions["A"].width = 26
    notes_sheet.column_dimensions["B"].width = 120

    for sheet in (summary_sheet, detail_sheet):
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column in (2, 3, 4, 11, 12, 13):
                    cell.number_format = '#,##0.00'
        sheet.auto_filter.ref = sheet.dimensions

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(report_data["summary"]), output_path


def summary_headers():
    return [
        "SOURCES_OF_COLLECTIONS",
        "TOTAL_COLLECTIONS",
        "NATIONAL",
        "PROVINCIAL_GENERAL_FUND",
        "PROVINCIAL_SEF",
        "PROVINCIAL_TOTAL",
        "MUNICIPAL_GENERAL_FUND",
        "MUNICIPAL_SEF",
        "MUNICIPAL_TRUST_FUND",
        "MUNICIPAL_TOTAL",
        "BARANGAY_SHARE",
        "FISHERIES",
    ]


def add_amount(target, name, amount):
    target[name] = target.get(name, 0) + (amount or 0)


def classify_summary_source(itaxtype, source_id, source_ct):
    code = (itaxtype or "").strip()
    source_id = int(source_id) if source_id is not None else None
    if source_ct in ("CTCI", "CTCC") or code == "CTC":
        return "Community Tax"
    if code == "MAS":
        return "Manufacturing"
    if code == "WHO":
        return "Distributor"
    if code == "RET":
        return "Retailing"
    if code == "BFI":
        return "Banks & Other Financial Int."
    if code in ("OBT", "CIC", "PED", "EMD") and source_id not in (807, 808):
        return "Other Business Tax"
    if code in ("TSG", "TSB"):
        return "Sand & Gravel"
    if code == "FPT":
        return "Fines & Penalties"
    if code == "MP":
        return "Mayor's Permit"
    if code == "FWM":
        return "Weights & Measures"
    if code in ("TOP", "MTO", "FLF"):
        return "Tricycle Permit Fee"
    if code == "OCC":
        return "Occupation Tax"
    if code == "COO":
        return "Cert. of Ownership"
    if code == "COT":
        return "Cert. of Transfer"
    if code == "CS":
        return "Cockpit Share"
    if code == "FRF" and source_id in (580, 639):
        return "Docking and Mooring Fee"
    if code in ("ST", "ATM"):
        return "Sultadas"
    if code in ("IM", "OPF", "SBF") or source_id in (807, 808):
        return "Miscellaneous"
    if code == "RB":
        return "Registration of Birth"
    if code == "RM":
        return "Marriage Fees"
    if code == "BF":
        return "Burial Fees"
    if code == "CE":
        return "Correction of Entry"
    if code in ("FRF", "IF"):
        return "Fishing Permit Fee"
    if code == "IAP":
        return "Sale of Agri. Prod."
    if code == "IAF":
        return "Sale of Acc. Forms"
    if code in ("WTR", "IWO"):
        return "Water Fees"
    if code in ("RFM", "MSF"):
        return "Market Stall Fee"
    if code == "SPF" or code == "RFS":
        return "Slaughterhouse Fee"
    if code in ("RFR", "IPG", "ICO"):
        return "Rent of Equipment"
    if code == "SF" and source_id == 810:
        return "Doc Stamp Tax"
    if code in ("SF", "PCL", "HEC", "OCL"):
        return "Secretary Fees"
    if code == "MDL":
        return "Med./Lab. Fees"
    if code == "GCF":
        return "Garbage Fees"
    if code in ("PFB", "BUF", "INS"):
        return "Building Permit Fee"
    if code == "EP":
        return "Electrical Permit Fee"
    if code == "ZLC":
        return "Zoning Fee"
    if code == "IFL":
        return "Livestock"
    if code == "IFD":
        return "Diving Fee"
    return None


GENERAL_ABSTRACT_COLUMNS = {
    "Manufacturing": 4,
    "Distributor": 5,
    "Retailing": 6,
    "Banks & Other Financial Int.": 7,
    "Other Business Tax": 8,
    "Sand & Gravel": 9,
    "Fines & Penalties": 10,
    "Mayor's Permit": 11,
    "Weights & Measures": 12,
    "Tricycle Permit Fee": 13,
    "Occupation Tax": 14,
    "Cert. of Ownership": 15,
    "Cert. of Transfer": 16,
    "Docking and Mooring Fee": 19,
    "Sultadas": 20,
    "Miscellaneous": 21,
    "Registration of Birth": 22,
    "Marriage Fees": 23,
    "Burial Fees": 24,
    "Correction of Entry": 25,
    "Fishing Permit Fee": 26,
    "Sale of Agri. Prod.": 27,
    "Sale of Acc. Forms": 28,
    "Water Fees": 29,
    "Market Stall Fee": 30,
    "Slaughterhouse Fee": 32,
    "Rent of Equipment": 33,
    "Doc Stamp Tax": 34,
    "Police Clearance": 35,
    "Secretary Fees": 36,
    "Med./Lab. Fees": 37,
    "Garbage Fees": 38,
}


TRUST_ABSTRACT_NAMES = {
    "Building Permit Fee",
    "Electrical Permit Fee",
    "Zoning Fee",
    "Livestock",
    "Diving Fee",
}


def split_summary_amount(name, amount):
    amount = amount or 0
    row = [name] + [0] * 11
    if name == "Cockpit Share":
        row[1] = amount
        row[3] = amount * Decimal("0.50")
        row[5] = row[3]
        row[6] = amount * Decimal("0.50")
        row[9] = row[6]
    elif name == "Building Permit Fee":
        row[1] = amount
        row[2] = amount * Decimal("0.05")
        row[6] = amount * Decimal("0.80")
        row[8] = amount * Decimal("0.15")
        row[9] = row[6] + row[8]
    elif name == "Livestock":
        row[1] = amount
        row[2] = amount * Decimal("0.20")
        row[6] = amount * Decimal("0.80")
        row[9] = row[6]
    elif name == "Diving Fee":
        row[1] = amount
        row[6] = amount * Decimal("0.40")
        row[9] = row[6]
        row[10] = amount * Decimal("0.30")
        row[11] = amount * Decimal("0.30")
    else:
        row[1] = amount
        row[6] = amount
        row[9] = amount
    return row


def open_firebird_connection(user="SYSDBA", password="masterkey", charset="UTF8"):
    fb_client = resolve_fb_client()
    last_error = None
    for db_path in db_path_candidates():
        db_path = (db_path or "").strip()
        if not db_path:
            continue
        if not is_firebird_server_path(db_path) and not Path(db_path).exists():
            print(f"STATUS: Skipping missing local database path: {db_path}")
            continue
        print(f"STATUS: Connecting to Firebird database: {db_path}")
        print(f"STATUS: Firebird client: {fb_client}")
        try:
            connection = fdb.connect(
                dsn=db_path,
                user=user,
                password=password,
                charset=charset,
                fb_library_name=fb_client,
                isolation_level=fdb.ISOLATION_LEVEL_READ_COMMITED_RO,
                no_db_triggers=True,
                no_gc=True,
            )
        except Exception as exc:
            last_error = exc
            print(f"STATUS: Database connection failed: {exc}")
            continue
        print("STATUS: Database connected successfully.")
        return connection
    if last_error is not None:
        raise last_error
    raise RuntimeError("No usable Firebird database path was found.")

def open_odbc_connection(user="SYSDBA", password="masterkey", charset="UTF8"):
    dsn = resolve_odbc_dsn()
    fb_client = resolve_fb_client()
    print(f"STATUS: Connecting through ODBC DSN: {dsn}")
    print(f"STATUS: ODBC Firebird client: {fb_client}")
    pyodbc = ensure_python_package("pyodbc")
    connection_string = f"DSN={dsn};UID={user};PWD={password};CLIENT={fb_client};"
    try:
        connection = pyodbc.connect(connection_string, autocommit=False)
    except Exception as exc:
        print(f"STATUS: ODBC connection failed: {exc}")
        raise
    print("STATUS: ODBC connected successfully.")
    return connection

def open_database_connection(user="SYSDBA", password="masterkey", charset="UTF8"):
    mode = (CONNECTION_MODE or "odbc").strip().lower()
    if mode == "odbc":
        return open_odbc_connection(user=user, password=password, charset=charset)
    if mode == "native":
        return open_firebird_connection(user=user, password=password, charset=charset)
    if mode == "auto":
        try:
            return open_firebird_connection(user=user, password=password, charset=charset)
        except Exception as native_exc:
            print(f"STATUS: Native Firebird connection unavailable; trying ODBC. Native error: {native_exc}")
            return open_odbc_connection(user=user, password=password, charset=charset)
    raise RuntimeError(f"Unsupported connection mode: {mode}")

def connect_report_db(user="SYSDBA", password="masterkey", charset="UTF8"):
    return open_database_connection(user=user, password=password, charset=charset)


def normalize_collector_name(collector):
    value = (collector or "").strip()
    if not value:
        return ""
    return COLLECTOR_ALIASES.get(value.lower(), value)


def payment_status_label(status_ct, void_bv):
    status = (status_ct or "").strip().upper()
    if void_bv:
        return "Void"
    if status in ("VOID", "VOI"):
        return "Void"
    if status in ("CNL", "CAN", "CNC", "CANCEL", "CANCELLED"):
        return "Cancelled"
    return "Paid"


def collector_choices_for_period(date_from, date_to, user, password):
    sql = """
        SELECT
            COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), NULLIF(TRIM(p.USERID), ''), 'UNSPECIFIED') AS COLLECTOR_NAME,
            COUNT(*) AS RECEIPTS,
            SUM(
                CASE
                    WHEN COALESCE(p.PAYGROUP_CT, '') = 'RPT'
                        THEN COALESCE(rpt_totals.RPT_TOTAL, detail_totals.DETAIL_TOTAL, p.AMOUNT, 0)
                    ELSE COALESCE(detail_totals.DETAIL_TOTAL, rpt_totals.RPT_TOTAL, p.AMOUNT, 0)
                END
            ) AS TOTAL_AMOUNT
        FROM PAYMENT p
        LEFT JOIN (
            SELECT PAYMENT_ID, SUM(AMOUNTPAID) AS DETAIL_TOTAL
            FROM PAYMENTDETAIL
            GROUP BY PAYMENT_ID
        ) detail_totals ON detail_totals.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN (
            SELECT PAYMENT_ID, SUM(AMOUNT) AS RPT_TOTAL
            FROM PAYMENTCLASSDETAIL
            GROUP BY PAYMENT_ID
        ) rpt_totals ON rpt_totals.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
        GROUP BY 1
        ORDER BY 1
    """
    con = connect_report_db(user, password, charset="WIN1252")
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        choices = [
            {"collector": collector_name, "receipts": receipts or 0, "total": total_amount or Decimal("0")}
            for collector_name, receipts, total_amount in cur.fetchall()
        ]
        con.rollback()
    finally:
        con.close()
    return choices


def resolve_collector_selection(collector_arg, date_from, date_to, user, password):
    value = (collector_arg or "").strip()
    if value and not value.isdigit():
        return normalize_collector_name(value)

    choices = collector_choices_for_period(date_from, date_to, user, password)
    if not choices:
        raise RuntimeError(f"No collectors found from {date_from} to {date_to}.")

    print(f"\nCollectors from {date_from} to {date_to}:")
    for index, choice in enumerate(choices, start=1):
        total = excel_value(choice["total"])
        print(f"{index:>2}. {choice['collector']:<24} {choice['receipts']:>6} receipts   {total:,.2f}")

    if value:
        selected_index = int(value)
    else:
        selected_value = input("Choose collector number: ").strip()
        if not selected_value:
            raise RuntimeError("Collector selection is required for report 34.")
        if not selected_value.isdigit():
            return normalize_collector_name(selected_value)
        selected_index = int(selected_value)

    if selected_index < 1 or selected_index > len(choices):
        raise RuntimeError(f"Collector number {selected_index} is outside the available range 1-{len(choices)}.")

    selected_collector = choices[selected_index - 1]["collector"]
    print(f"Selected collector: {selected_collector}")
    return selected_collector


def build_no_rpt_rows_from_fdb(date_from, date_to, user, password):
    order = [
        "Manufacturing", "Distributor", "Retailing", "Banks & Other Financial Int.",
        "Other Business Tax", "Sand & Gravel", "Fines & Penalties", "Mayor's Permit",
        "Weights & Measures", "Tricycle Permit Fee", "Occupation Tax", "Cert. of Ownership",
        "Cert. of Transfer", "Cockpit Share", "Docking and Mooring Fee", "Sultadas",
        "Miscellaneous", "Registration of Birth", "Marriage Fees", "Burial Fees",
        "Correction of Entry", "Fishing Permit Fee", "Sale of Agri. Prod.",
        "Sale of Acc. Forms", "Water Fees", "Market Stall Fee", "Cash Tickets",
        "Slaughterhouse Fee", "Rent of Equipment", "Doc Stamp Tax", "Secretary Fees",
        "Med./Lab. Fees", "Garbage Fees", "Cutting Tree", "Community Tax",
        "Building Permit Fee", "Electrical Permit Fee", "Zoning Fee", "Livestock",
        "Diving Fee",
    ]
    amounts = {name: 0 for name in order}
    sql = """
        SELECT
            pd.ITAXTYPE_CT,
            pd.SOURCEID,
            pd.SOURCE_CT,
            SUM(pd.AMOUNTPAID) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(p.PAYGROUP_CT, '') <> 'RPT'
        GROUP BY pd.ITAXTYPE_CT, pd.SOURCEID, pd.SOURCE_CT
    """
    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for itaxtype, source_id, source_ct, amount in cur.fetchall():
            name = classify_summary_source(
                itaxtype.strip() if isinstance(itaxtype, str) else itaxtype,
                source_id,
                source_ct.strip() if isinstance(source_ct, str) else source_ct,
            )
            if name:
                add_amount(amounts, name, amount)
        con.rollback()
    finally:
        con.close()

    tax_business_amounts = tax_on_business_summary_amounts(date_from, date_to)
    if any(tax_business_amounts.values()):
        for name, amount in tax_business_amounts.items():
            if name in amounts:
                amounts[name] = amount

    rows = [summary_headers()]
    rows.extend(split_summary_amount(name, amounts.get(name, 0)) for name in order)
    totals = ["TOTAL"]
    for i in range(1, 12):
        totals.append(sum(row[i] for row in rows[1:] if isinstance(row[i], (int, float, Decimal))))
    rows.append(totals)
    return rows


def rpt_summary_rows_from_fdb(date_from, date_to, user, password):
    sql = """
        SELECT
            pcd.PROPERTYKIND_CT,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        GROUP BY pcd.PROPERTYKIND_CT, pcd.ITAXTYPE_CT, pcd.CASETYPE_CT, pcd.TAXYEAR
    """
    con = connect_report_db(user, password)
    entries = []
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for propkind, itaxtype, casetype, taxyear, amount in cur.fetchall():
            entries.append({
                "property_group": "Land" if (propkind or "").strip() == "L" else "Bldg.",
                "tax_type": (itaxtype or "").strip(),
                "case_type": (casetype or "").strip(),
                "taxyear": taxyear,
                "amount": amount or 0,
            })
        con.rollback()
    finally:
        con.close()

    current_taxyear = max((e["taxyear"] for e in entries if e["taxyear"] is not None), default=None)
    buckets = {}
    for e in entries:
        if e["case_type"] == "PEN":
            line = "Penalties"
        elif e["taxyear"] == current_taxyear:
            line = "Current Year"
        else:
            line = "Previous Years"
        key = (e["property_group"], e["tax_type"], line)
        buckets[key] = buckets.get(key, 0) + e["amount"]
    return buckets, current_taxyear


def rpt_row(label, amount, tax_type):
    row = [label] + [0] * 11
    amount = amount or 0
    row[1] = amount
    if tax_type == "BSC":
        row[3] = amount * Decimal("0.35")
        row[5] = row[3]
        row[6] = amount * Decimal("0.40")
        row[9] = row[6]
        row[10] = amount * Decimal("0.25")
    else:
        row[4] = amount * Decimal("0.50")
        row[5] = row[4]
        row[7] = amount * Decimal("0.50")
        row[9] = row[7]
    return row


def build_rpt_rows_from_fdb(date_from, date_to, user, password):
    buckets, _ = rpt_summary_rows_from_fdb(date_from, date_to, user, password)
    rows = [summary_headers()]
    layout = [
        ("Real Property Tax - Basic/Land", None, None),
        ("Current Year", "Land", "BSC"),
        ("Previous Years", "Land", "BSC"),
        ("Penalties", "Land", "BSC"),
        ("Real Property Tax - SEF/Land", None, None),
        ("Current Year", "Land", "SEF"),
        ("Previous Years", "Land", "SEF"),
        ("Penalties", "Land", "SEF"),
        ("Real Property Tax - Basic/Bldg.", None, None),
        ("Current Year", "Bldg.", "BSC"),
        ("Previous Years", "Bldg.", "BSC"),
        ("Penalties", "Bldg.", "BSC"),
        ("Real Property Tax - SEF/Bldg.", None, None),
        ("Current Year", "Bldg.", "SEF"),
        ("Previous Years", "Bldg.", "SEF"),
        ("Penalties", "Bldg.", "SEF"),
    ]
    for label, group, tax_type in layout:
        if group is None:
            rows.append([label] + [""] * 11)
        else:
            rows.append(rpt_row(label, buckets.get((group, tax_type, label), 0), tax_type))
    totals = ["TOTAL"]
    for i in range(1, 12):
        totals.append(sum(row[i] for row in rows[1:] if isinstance(row[i], (int, float, Decimal))))
    rows.append(totals)
    return rows


def build_full_summary_rows_from_fdb(date_from, date_to, user, password):
    no_rpt = build_no_rpt_rows_from_fdb(date_from, date_to, user, password)
    rpt = build_rpt_rows_from_fdb(date_from, date_to, user, password)
    rows = [summary_headers()]
    rows.extend(no_rpt[1:-1])
    rows.extend(rpt[1:-1])
    totals = ["TOTAL"]
    for i in range(1, 12):
        totals.append(sum(row[i] for row in rows[1:] if isinstance(row[i], (int, float, Decimal))))
    rows.append(totals)
    return rows


def build_rpt_detail_summary_from_fdb(date_from, date_to, user, password):
    buckets, current_taxyear = rpt_summary_rows_from_fdb(date_from, date_to, user, password)
    rows = [[
        "PROPERTY_GROUP", "TAX_TYPE", "LINE_TYPE", "CURRENT_TAXYEAR_USED",
        "FDB_AMOUNT", "PROVINCIAL_SHARE", "MUNICIPAL_SHARE",
        "BARANGAY_SHARE", "SUMMARY_TEMPLATE_SECTION",
    ]]
    for group in ("Land", "Bldg."):
        for tax_type in ("BSC", "SEF"):
            for line in ("Current Year", "Previous Years", "Penalties"):
                amount = buckets.get((group, tax_type, line), 0)
                if tax_type == "BSC":
                    prov = amount * Decimal("0.35")
                    mun = amount * Decimal("0.40")
                    brgy = amount * Decimal("0.25")
                else:
                    prov = amount * Decimal("0.50")
                    mun = amount * Decimal("0.50")
                    brgy = 0
                section = f"Real Property Tax - {'Basic' if tax_type == 'BSC' else 'SEF'}/{group}"
                rows.append([group, tax_type, line, current_taxyear, amount, prov, mun, brgy, section])
    return rows


def rpt_record_headers():
    return [
        "DATE",
        "PAID_BY",
        "NAME_OF_TAXPAYER",
        "PERIOD_COVERED",
        "PIN",
        "OR_NO",
        "TD_ARP_NO",
        "BARANGAY",
        "BASIC_CURRENT_YEAR_GROSS",
        "BASIC_DISCOUNT",
        "BASIC_PRIOR_YEARS",
        "BASIC_PENALTY_CURRENT_YEAR",
        "BASIC_PENALTY_PREV_YEARS",
        "BASIC_PENALTY_PRIOR_YEARS",
        "BASIC_GROSS",
        "BASIC_NET",
        "SEF_CURRENT_YEAR_GROSS",
        "SEF_DISCOUNT",
        "SEF_PRIOR_YEARS",
        "SEF_PENALTY_CURRENT_YEAR",
        "SEF_PENALTY_PREV_YEARS",
        "SEF_PENALTY_PRIOR_YEARS",
        "SEF_GROSS",
        "SEF_NET",
        "GRAND_GROSS",
        "GRAND_NET",
        "BASIC_25_PERCENT_SHARE",
        "PROPERTY_CLASSIFICATION",
        "PROPERTY_KIND",
        "COLLECTOR",
        "PAYMENT_STATUS_CT",
        "IS_CANCELLED",
        "PAYMENT_TOTAL_AMOUNT",
        "BOOKINGREFERENCE",
        "IS_VOID",
        "INCLUDE_IN_REPORT",
    ]


def period_covered(taxyears):
    years = sorted(year for year in taxyears if year is not None)
    if not years:
        return ""
    if len(years) == 1:
        return str(years[0])
    if years == list(range(years[0], years[-1] + 1)):
        return f"{years[0]}-{years[-1]}"
    return ", ".join(str(year) for year in years)


def display_lookup(description, code):
    description = (description or "").strip()
    code = (code or "").strip()
    if description:
        return description
    return code


def property_classification_label(description, code):
    code = (code or "").strip()
    if code.upper().startswith("S"):
        return "SPECIAL"
    return display_lookup(description, code)


def property_classification_key(code):
    code = (code or "").strip().upper()
    if code.startswith("S"):
        return "SPECIAL"
    return code


def add_rpt_record_amount(record, tax_type, case_type, taxyear, amount, current_taxyear):
    tax_prefix = "basic" if tax_type == "BSC" else "sef" if tax_type == "SEF" else None
    if tax_prefix is None:
        return

    amount = amount or 0
    if case_type == "DED":
        record[f"{tax_prefix}_discount"] += abs(amount)
    elif case_type == "PEN":
        if taxyear == current_taxyear:
            record[f"{tax_prefix}_pen_current"] += amount
        elif current_taxyear is not None and taxyear == current_taxyear - 1:
            record[f"{tax_prefix}_pen_prev"] += amount
        else:
            record[f"{tax_prefix}_pen_prior"] += amount
    else:
        if taxyear == current_taxyear:
            record[f"{tax_prefix}_current"] += amount
        else:
            record[f"{tax_prefix}_prior"] += amount


def build_rpt_record_rows_from_fdb(date_from, date_to, user, password):
    _, current_taxyear = rpt_summary_rows_from_fdb(date_from, date_to, user, password)
    sql = """
        SELECT
            p.PAYMENT_ID,
            p.PAYMENTDATE,
            p.PAIDBY,
            tx.OWNERNAME,
            p.RECEIPTNO,
            p.STATUS_CT,
            p.AMOUNT,
            p.RCDNUMBER,
            p.VOID_BV,
            COALESCE(p.COLLECTOR, p.USERID) AS COLLECTOR_NAME,
            pcd.TAXTRANS_ID,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            pcd.AMOUNT AS DETAIL_AMOUNT,
            pcd.CANCELLED_BV,
            pcd.CLASSCODE_CT,
            pcd.PROPERTYKIND_CT,
            ra.TDNO,
            ra.TDNOFORGR,
            ra.PREDOMCLASSCODE_CT,
            prop.PINNO,
            prop.NEWPINNO,
            brgy.DESCRIPTION AS BARANGAY_NAME,
            cls.DESCRIPTION AS CLASSIFICATION_NAME,
            kind.DESCRIPTION AS PROPERTY_KIND_NAME
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN TAXPAYER tx ON tx.LOCAL_TIN = p.LOCAL_TIN
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        LEFT JOIN PROPERTY prop ON prop.PROP_ID = ra.PROP_ID
        LEFT JOIN T_BARANGAY brgy
               ON brgy.CODE = prop.BARANGAY_CT
              AND brgy.MUNICIPAL_ID = prop.MUNICIPAL_ID
              AND brgy.PROVINCE_CT = prop.PROVINCE_CT
        LEFT JOIN T_CLASSIFICATION cls
               ON cls.CODE = COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT)
        LEFT JOIN T_PROPERTYKIND kind
               ON kind.CODE = COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT)
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        ORDER BY p.PAYMENTDATE, p.RECEIPTNO, p.PAYMENT_ID, pcd.TAXTRANS_ID, pcd.TAXYEAR
    """
    amount_fields = (
        "basic_current", "basic_discount", "basic_prior",
        "basic_pen_current", "basic_pen_prev", "basic_pen_prior",
        "sef_current", "sef_discount", "sef_prior",
        "sef_pen_current", "sef_pen_prev", "sef_pen_prior",
    )
    records = {}
    order = []

    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for (
            payment_id, payment_date, paid_by, taxpayer_name, receipt_no, status_ct,
            payment_amount, booking_reference, void_bv, collector_name, taxtrans_id,
            tax_type, case_type, taxyear, detail_amount, cancelled_bv, class_code,
            property_kind_code, td_no, td_no_for_gr, predom_class_code, pin_no,
            new_pin_no, barangay_name, classification_name, property_kind_name,
        ) in cur.fetchall():
            classification_code = class_code or predom_class_code
            key = (payment_id, taxtrans_id, property_classification_key(classification_code))
            if key not in records:
                records[key] = {
                    "payment_date": payment_date,
                    "paid_by": paid_by,
                    "taxpayer_name": taxpayer_name or paid_by,
                    "taxyears": set(),
                    "pin": new_pin_no or pin_no,
                    "receipt_no": receipt_no,
                    "td_arp_no": td_no_for_gr or td_no,
                    "barangay": barangay_name,
                    "classification": property_classification_label(
                        classification_name,
                        classification_code,
                    ),
                    "property_kind": display_lookup(property_kind_name, property_kind_code),
                    "collector": collector_name,
                    "status_ct": status_ct,
                    "is_cancelled": cancelled_bv or 0,
                    "payment_amount": payment_amount or 0,
                    "booking_reference": booking_reference,
                    "is_void": void_bv or 0,
                    "include_in_report": 1,
                }
                for field in amount_fields:
                    records[key][field] = Decimal("0")
                order.append(key)

            record = records[key]
            record["taxyears"].add(taxyear)
            if not record["classification"]:
                record["classification"] = property_classification_label(
                    classification_name,
                    classification_code,
                )
            if not record["property_kind"]:
                record["property_kind"] = display_lookup(property_kind_name, property_kind_code)
            add_rpt_record_amount(
                record,
                (tax_type or "").strip(),
                (case_type or "").strip(),
                taxyear,
                detail_amount,
                current_taxyear,
            )
        con.rollback()
    finally:
        con.close()

    rows = [rpt_record_headers()]
    for key in order:
        record = records[key]
        basic_gross = (
            record["basic_current"] + record["basic_prior"] +
            record["basic_pen_current"] + record["basic_pen_prev"] + record["basic_pen_prior"]
        )
        basic_net = basic_gross - record["basic_discount"]
        sef_gross = (
            record["sef_current"] + record["sef_prior"] +
            record["sef_pen_current"] + record["sef_pen_prev"] + record["sef_pen_prior"]
        )
        sef_net = sef_gross - record["sef_discount"]
        grand_gross = basic_gross + sef_gross
        grand_net = basic_net + sef_net
        rows.append([
            record["payment_date"],
            record["paid_by"],
            record["taxpayer_name"],
            period_covered(record["taxyears"]),
            record["pin"],
            record["receipt_no"],
            record["td_arp_no"],
            record["barangay"],
            record["basic_current"],
            record["basic_discount"],
            record["basic_prior"],
            record["basic_pen_current"],
            record["basic_pen_prev"],
            record["basic_pen_prior"],
            basic_gross,
            basic_net,
            record["sef_current"],
            record["sef_discount"],
            record["sef_prior"],
            record["sef_pen_current"],
            record["sef_pen_prev"],
            record["sef_pen_prior"],
            sef_gross,
            sef_net,
            grand_gross,
            grand_net,
            basic_net * Decimal("0.25"),
            record["classification"],
            record["property_kind"],
            record["collector"],
            record["status_ct"],
            record["is_cancelled"],
            record["payment_amount"],
            record["booking_reference"],
            record["is_void"],
            record["include_in_report"],
        ])
    return rows


def report_year_from_date(date_from):
    return datetime.strptime(date_from, "%Y-%m-%d").year


def build_advance_rpt_record_rows_from_fdb(date_from, date_to, user, password):
    report_year = report_year_from_date(date_from)
    sql = """
        SELECT
            p.PAYMENT_ID,
            p.PAYMENTDATE,
            p.PAIDBY,
            tx.OWNERNAME,
            p.RECEIPTNO,
            COALESCE(p.COLLECTOR, p.USERID) AS COLLECTOR_NAME,
            pcd.TAXTRANS_ID,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            pcd.AMOUNT AS DETAIL_AMOUNT,
            pcd.CLASSCODE_CT,
            pcd.PROPERTYKIND_CT,
            ra.TDNO,
            ra.TDNOFORGR,
            ra.PREDOMCLASSCODE_CT,
            prop.PINNO,
            prop.NEWPINNO,
            brgy.DESCRIPTION AS BARANGAY_NAME,
            cls.DESCRIPTION AS CLASSIFICATION_NAME,
            kind.DESCRIPTION AS PROPERTY_KIND_NAME
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN TAXPAYER tx ON tx.LOCAL_TIN = p.LOCAL_TIN
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        LEFT JOIN PROPERTY prop ON prop.PROP_ID = ra.PROP_ID
        LEFT JOIN T_BARANGAY brgy
               ON brgy.CODE = prop.BARANGAY_CT
              AND brgy.MUNICIPAL_ID = prop.MUNICIPAL_ID
              AND brgy.PROVINCE_CT = prop.PROVINCE_CT
        LEFT JOIN T_CLASSIFICATION cls
               ON cls.CODE = COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT)
        LEFT JOIN T_PROPERTYKIND kind
               ON kind.CODE = COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT)
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
          AND pcd.TAXYEAR > ?
        ORDER BY p.PAYMENTDATE, p.RECEIPTNO, p.PAYMENT_ID, pcd.TAXTRANS_ID, pcd.TAXYEAR
    """
    records = {}
    order = []

    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to, report_year))
        for (
            payment_id, payment_date, paid_by, taxpayer_name, receipt_no,
            collector_name, taxtrans_id, tax_type, case_type, taxyear, detail_amount,
            class_code, property_kind_code, td_no, td_no_for_gr, predom_class_code,
            pin_no, new_pin_no, barangay_name, classification_name, property_kind_name,
        ) in cur.fetchall():
            classification_code = class_code or predom_class_code
            key = (payment_id, taxtrans_id, property_classification_key(classification_code))
            if key not in records:
                records[key] = {
                    "payment_date": payment_date,
                    "paid_by": paid_by,
                    "taxpayer_name": taxpayer_name or paid_by,
                    "taxyears": set(),
                    "pin": new_pin_no or pin_no,
                    "receipt_no": receipt_no,
                    "td_arp_no": td_no_for_gr or td_no,
                    "barangay": barangay_name,
                    "basic_gross": Decimal("0"),
                    "basic_discount": Decimal("0"),
                    "sef_gross": Decimal("0"),
                    "sef_discount": Decimal("0"),
                    "classification": property_classification_label(
                        classification_name,
                        classification_code,
                    ),
                    "property_kind": display_lookup(property_kind_name, property_kind_code),
                    "collector": collector_name,
                }
                order.append(key)

            record = records[key]
            record["taxyears"].add(taxyear)
            prefix = "basic" if (tax_type or "").strip() == "BSC" else (
                "sef" if (tax_type or "").strip() == "SEF" else None
            )
            if prefix is None:
                continue
            if (case_type or "").strip() == "DED":
                record[f"{prefix}_discount"] += abs(detail_amount or 0)
            else:
                record[f"{prefix}_gross"] += detail_amount or 0
        con.rollback()
    finally:
        con.close()

    rows = [[
        "DATE", "PAID_BY", "NAME_OF_TAXPAYER", "PERIOD_COVERED", "PIN", "OR_NO",
        "TD_ARP_NO", "BARANGAY", "BSC_GROSS_AMOUNT", "BSC_DISCOUNT",
        "BSC_TOTAL_COLLECTION", "SEF_GROSS_AMOUNT", "SEF_DISCOUNT",
        "SEF_TOTAL_COLLECTION", "GRAND_TOTAL", "PROPERTY_CLASSIFICATION",
        "PROPERTY_KIND", "COLLECTOR",
    ]]
    for key in order:
        record = records[key]
        basic_net = record["basic_gross"] - record["basic_discount"]
        sef_net = record["sef_gross"] - record["sef_discount"]
        rows.append([
            record["payment_date"],
            record["paid_by"],
            record["taxpayer_name"],
            period_covered(record["taxyears"]),
            record["pin"],
            record["receipt_no"],
            record["td_arp_no"],
            record["barangay"],
            record["basic_gross"],
            record["basic_discount"],
            basic_net,
            record["sef_gross"],
            record["sef_discount"],
            sef_net,
            basic_net + sef_net,
            record["classification"],
            record["property_kind"],
            record["collector"],
        ])
    return rows


def sharing_row_for_classification(property_kind, class_code):
    property_kind = (property_kind or "").strip()
    class_code = (class_code or "").strip()
    if property_kind == "L":
        if class_code == "A":
            return 11
        if class_code == "R":
            return 12
        if class_code == "C":
            return 13
        return 14

    if property_kind == "M":
        return 22
    if class_code == "R":
        return 23
    if class_code == "C":
        return 24
    if class_code == "A":
        return 25
    if class_code.upper().startswith("S"):
        return 26
    return 26


def provincial_rpt_row_for_detail(property_kind, class_code, actual_use):
    property_kind = (property_kind or "").strip()
    class_code = (class_code or "").strip().upper()
    actual_use = (actual_use or "").strip().upper()

    if property_kind == "M":
        return 19

    is_land = property_kind == "L"
    if class_code.startswith("S") or actual_use == "ARC":
        return 14 if is_land else 21
    if class_code == "R" or actual_use in ("AR", "ARD"):
        return 9 if is_land else 16
    if class_code == "C" or actual_use == "AC":
        return 10 if is_land else 17
    if class_code == "I" or actual_use == "AI":
        return 11 if is_land else 18
    if class_code == "M" or actual_use == "AM":
        return 12 if is_land else 19
    if class_code == "A" or actual_use == "AA":
        return 13 if is_land else 20
    if class_code == "T" or actual_use == "ATF":
        return 15 if is_land else 21
    return 14 if is_land else 21


def build_provincial_rpt_coding_rows_from_fdb(date_from, date_to, user, password):
    _, current_taxyear = rpt_summary_rows_from_fdb(date_from, date_to, user, password)
    if current_taxyear is None:
        current_taxyear = report_year_from_date(date_from)

    sql = """
        SELECT
            COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT) AS PROPERTYKIND_CT,
            COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT) AS CLASSCODE_CT,
            pcd.ACTUALUSE_CT,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        LEFT JOIN PROPERTY prop ON prop.PROP_ID = ra.PROP_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
          AND pcd.ITAXTYPE_CT IN ('BSC', 'SEF')
        GROUP BY COALESCE(pcd.PROPERTYKIND_CT, prop.PROPERTYKIND_CT),
                 COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT),
                 pcd.ACTUALUSE_CT,
                 pcd.ITAXTYPE_CT,
                 pcd.CASETYPE_CT,
                 pcd.TAXYEAR
    """
    values = {}
    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for property_kind, class_code, actual_use, tax_type, case_type, taxyear, amount in cur.fetchall():
            tax_type = (tax_type or "").strip()
            case_type = (case_type or "").strip()
            sheet_name = "GF" if tax_type == "BSC" else "SEF" if tax_type == "SEF" else None
            if sheet_name is None:
                continue

            row_index = provincial_rpt_row_for_detail(property_kind, class_code, actual_use)
            if case_type == "PEN":
                col_index = 7 if taxyear == current_taxyear else 9
                value = amount or 0
            else:
                col_index = 3 if taxyear == current_taxyear else 5
                if case_type == "DED":
                    value = -abs(amount or 0)
                else:
                    value = amount or 0

            share = Decimal("0.35") if tax_type == "BSC" else Decimal("0.50")
            key = (sheet_name, row_index, col_index)
            values[key] = values.get(key, Decimal("0")) + (value * share)
        con.rollback()
    finally:
        con.close()

    rows = [["SHEET", "ROW", "COLUMN", "VALUE"]]
    for sheet_name in ("GF", "SEF"):
        for row_index in range(9, 22):
            for col_index in (3, 5, 7, 9):
                rows.append([sheet_name, row_index, col_index, values.get((sheet_name, row_index, col_index), Decimal("0"))])
    return rows


def payment_detail_rows_for_abstract(date_from, date_to, user, password):
    sql = """
        SELECT
            p.PAYMENT_ID,
            p.PAYMENTDATE,
            p.RECEIPTNO,
            p.PAIDBY,
            COALESCE(p.COLLECTOR, p.USERID) AS COLLECTOR_NAME,
            p.PAYGROUP_CT,
            pd.ITAXTYPE_CT,
            pd.SOURCEID,
            pd.SOURCE_CT,
            pd.AMOUNTPAID
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(p.PAYGROUP_CT, '') <> 'RPT'
        ORDER BY p.PAYMENTDATE, p.RECEIPTNO, p.PAYMENT_ID, pd.RECEIPTITEMORDER
    """
    rows = []
    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for row in cur.fetchall():
            rows.append(row)
        con.rollback()
    finally:
        con.close()
    return rows


def add_general_abstract_amount(record, source_name, amount):
    amount = amount or 0
    if source_name == "Cockpit Share":
        record[17] += amount * Decimal("0.50")
        record[18] += amount * Decimal("0.50")
        return True

    col_index = GENERAL_ABSTRACT_COLUMNS.get(source_name)
    if col_index is None:
        return False
    record[col_index] += amount
    return True


def build_abstract_general_collections_rows_from_fdb(date_from, date_to, user, password):
    records = {}
    order = []

    for (
        payment_id, payment_date, receipt_no, paid_by, collector_name, paygroup,
        itaxtype, source_id, source_ct, amount,
    ) in payment_detail_rows_for_abstract(date_from, date_to, user, password):
        source_name = classify_summary_source(
            itaxtype.strip() if isinstance(itaxtype, str) else itaxtype,
            source_id,
            source_ct.strip() if isinstance(source_ct, str) else source_ct,
        )
        if source_name in TRUST_ABSTRACT_NAMES or source_name == "Community Tax":
            continue

        if payment_id not in records:
            records[payment_id] = {
                "date": payment_date,
                "receipt_no": receipt_no,
                "paid_by": paid_by,
                "collector": collector_name,
                "paygroup": paygroup,
                "amounts": {col_index: Decimal("0") for col_index in range(4, 39)},
            }
            order.append(payment_id)

        added = add_general_abstract_amount(records[payment_id]["amounts"], source_name, amount)
        if not added and source_name:
            records[payment_id]["amounts"][21] += amount or 0

    headers = [
        "Date", "Receipt Number", "Names", "Manufacturing", "Distributor", "Retailing",
        "Financial", "Other", "Sand & Gravel", "Fines & Penalties", "Mayor's Permit",
        "W. & M.", "Trirycle Operators", "Occu.", "Cert. of Ownership", "Cert. of Transfer",
        "Cockpit Prov. Share", "Cockpit Local Share", "Docking and Mooring Fee", "Sultadas",
        "MISCS.", "Reg. of", "Marriage Fees", "Burial Fees", "Correction of Entry",
        "Fishing Permit Fee", "Sale of Agri. Prod.", "Sale of Acct. Form", "Water Fees",
        "Stall Fees", "Cash Tickets", "Slaughter House Fee", "Rental of Equipment",
        "Doc. Stamp", "Police Clearance", "Cert.", "Med./Dent. & Lab. Fees", "Garbage Fees",
        "CASHIER", "TOTAL", "TYPE OF RECIEPT",
    ]
    rows = [headers]
    daily = {}

    for payment_id in order:
        record = records[payment_id]
        total = sum(record["amounts"].values())
        if total == 0:
            continue
        row = [
            record["date"], record["receipt_no"], record["paid_by"],
            *[record["amounts"].get(col_index, Decimal("0")) for col_index in range(4, 39)],
            record["collector"], total, record["paygroup"],
        ]
        rows.append(row)

        day = record["date"].date() if hasattr(record["date"], "date") else record["date"]
        if day not in daily:
            daily[day] = {col_index: Decimal("0") for col_index in range(4, 39)}
        for col_index, value in record["amounts"].items():
            daily[day][col_index] += value

    daily_rows = [["Date"] + headers[3:38] + ["TOTAL"]]
    for day in sorted(daily):
        total = sum(daily[day].values())
        daily_rows.append([day] + [daily[day].get(col_index, Decimal("0")) for col_index in range(4, 39)] + [total])
    return rows, daily_rows


def trust_split_values(source_name, amount):
    amount = amount or 0
    values = {col_index: Decimal("0") for col_index in range(4, 14)}
    if source_name == "Building Permit Fee":
        values[4] = amount * Decimal("0.80")
        values[5] = amount * Decimal("0.15")
        values[6] = amount * Decimal("0.05")
    elif source_name == "Electrical Permit Fee":
        values[7] = amount
    elif source_name == "Zoning Fee":
        values[8] = amount
    elif source_name == "Livestock":
        values[9] = amount * Decimal("0.80")
        values[10] = amount * Decimal("0.20")
    elif source_name == "Diving Fee":
        values[11] = amount * Decimal("0.40")
        values[12] = amount * Decimal("0.30")
        values[13] = amount * Decimal("0.30")
    return values


def build_abstract_trust_funds_rows_from_fdb(date_from, date_to, user, password):
    records = {}
    order = []

    for (
        payment_id, payment_date, receipt_no, paid_by, collector_name, paygroup,
        itaxtype, source_id, source_ct, amount,
    ) in payment_detail_rows_for_abstract(date_from, date_to, user, password):
        source_name = classify_summary_source(
            itaxtype.strip() if isinstance(itaxtype, str) else itaxtype,
            source_id,
            source_ct.strip() if isinstance(source_ct, str) else source_ct,
        )
        if source_name not in TRUST_ABSTRACT_NAMES:
            continue

        if payment_id not in records:
            records[payment_id] = {
                "date": payment_date,
                "receipt_no": receipt_no,
                "paid_by": paid_by,
                "collector": collector_name,
                "paygroup": paygroup,
                "amounts": {col_index: Decimal("0") for col_index in range(4, 14)},
            }
            order.append(payment_id)

        split_values = trust_split_values(source_name, amount)
        for col_index, value in split_values.items():
            records[payment_id]["amounts"][col_index] += value

    headers = [
        "Date", "Receipt Number", "Names", "Building Fee 80% Local",
        "Building Fee 15% T.F.", "Building Fee 5% Nat'L.", "Electrical Fee",
        "Zoning Fee", "Livestock 80% Local", "Livestock 20% Nat'l",
        "Diving 40% GF", "Diving 30% Fishers", "Diving 30% Brgy",
        "CASHIER", "Total", "TYPE OF RECIEPT",
    ]
    rows = [headers]
    daily = {}

    for payment_id in order:
        record = records[payment_id]
        total = sum(record["amounts"].values())
        if total == 0:
            continue
        row = [
            record["date"], record["receipt_no"], record["paid_by"],
            *[record["amounts"].get(col_index, Decimal("0")) for col_index in range(4, 14)],
            record["collector"], total, record["paygroup"],
        ]
        rows.append(row)

        day = record["date"].date() if hasattr(record["date"], "date") else record["date"]
        if day not in daily:
            daily[day] = {col_index: Decimal("0") for col_index in range(4, 14)}
        for col_index, value in record["amounts"].items():
            daily[day][col_index] += value

    daily_rows = [["DATE"] + headers[3:13] + ["Total"]]
    for day in sorted(daily):
        total = sum(daily[day].values())
        daily_rows.append([day] + [daily[day].get(col_index, Decimal("0")) for col_index in range(4, 14)] + [total])
    return rows, daily_rows


def add_daily_amount(daily, day, column_name, amount):
    if day not in daily:
        daily[day] = {
            "CTC": Decimal("0"),
            "RPT": Decimal("0"),
            "GF_TF": Decimal("0"),
        }
    daily[day][column_name] += amount or 0


def build_full_report_collections_rows_from_fdb(date_from, date_to, user, password):
    daily = {}
    ctc_sql = """
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS COLLECTION_DATE,
            SUM(pd.AMOUNTPAID) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND (pd.SOURCE_CT IN ('CTCI', 'CTCC') OR pd.ITAXTYPE_CT = 'CTC')
        GROUP BY CAST(p.PAYMENTDATE AS DATE)
    """
    rpt_sql = """
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS COLLECTION_DATE,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        GROUP BY CAST(p.PAYMENTDATE AS DATE)
    """
    gf_tf_sql = """
        SELECT
            CAST(p.PAYMENTDATE AS DATE) AS COLLECTION_DATE,
            SUM(pd.AMOUNTPAID) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTDETAIL pd ON pd.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(p.PAYGROUP_CT, '') <> 'RPT'
          AND NOT (pd.SOURCE_CT IN ('CTCI', 'CTCC') OR pd.ITAXTYPE_CT = 'CTC')
        GROUP BY CAST(p.PAYMENTDATE AS DATE)
    """

    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        for sql, column_name in (
            (ctc_sql, "CTC"),
            (rpt_sql, "RPT"),
            (gf_tf_sql, "GF_TF"),
        ):
            cur.execute(sql, (date_from, date_to))
            for day, amount in cur.fetchall():
                add_daily_amount(daily, day, column_name, amount)
        con.rollback()
    finally:
        con.close()

    rows = [["DATE", "CTC", "RPT", "GF AND TF"]]
    for day in sorted(daily):
        rows.append([
            day,
            daily[day]["CTC"],
            daily[day]["RPT"],
            daily[day]["GF_TF"],
        ])
    return rows


def build_summary_sharing_rows_from_fdb(date_from, date_to, user, password):
    buckets, current_taxyear = rpt_summary_rows_from_fdb(date_from, date_to, user, password)
    if current_taxyear is None:
        current_taxyear = report_year_from_date(date_from)

    sql = """
        SELECT
            pcd.PROPERTYKIND_CT,
            COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT) AS CLASSCODE_CT,
            pcd.ITAXTYPE_CT,
            pcd.CASETYPE_CT,
            pcd.TAXYEAR,
            SUM(pcd.AMOUNT) AS AMOUNT
        FROM PAYMENT p
        JOIN PAYMENTCLASSDETAIL pcd ON pcd.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN RPTASSESSMENT ra ON ra.TAXTRANS_ID = pcd.TAXTRANS_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND p.PAYGROUP_CT = 'RPT'
          AND COALESCE(p.VOID_BV, 0) = 0
          AND COALESCE(TRIM(p.STATUS_CT), '') NOT IN ('CNL', 'CAN', 'CNC', 'CANCEL', 'CANCELLED', 'VOID', 'VOI')
          AND COALESCE(pcd.CANCELLED_BV, 0) = 0
        GROUP BY pcd.PROPERTYKIND_CT, COALESCE(pcd.CLASSCODE_CT, ra.PREDOMCLASSCODE_CT),
                 pcd.ITAXTYPE_CT, pcd.CASETYPE_CT, pcd.TAXYEAR
    """
    values = {}

    con = connect_report_db(user, password)
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to))
        for property_kind, class_code, tax_type, case_type, taxyear, amount in cur.fetchall():
            row = sharing_row_for_classification(property_kind, class_code)
            tax_type = (tax_type or "").strip()
            case_type = (case_type or "").strip()
            if tax_type == "BSC":
                current_col, discount_col, prior_col, pen_current_col, pen_prior_col = 3, 4, 5, 6, 7
            elif tax_type == "SEF":
                current_col, discount_col, prior_col, pen_current_col, pen_prior_col = 10, 11, 12, 13, 14
            else:
                continue

            if case_type == "DED":
                col = discount_col
                value = abs(amount or 0)
            elif case_type == "PEN":
                col = pen_current_col if taxyear == current_taxyear else pen_prior_col
                value = amount or 0
            else:
                col = current_col if taxyear == current_taxyear else prior_col
                value = amount or 0
            values[(row, col)] = values.get((row, col), Decimal("0")) + value
        con.rollback()
    finally:
        con.close()

    rows = [["ROW", "COLUMN", "VALUE"]]
    for row in (11, 12, 13, 14, 22, 23, 24, 25, 26):
        for col in (3, 4, 5, 6, 7, 10, 11, 12, 13, 14):
            rows.append([row, col, values.get((row, col), Decimal("0"))])
    return rows


def build_generate_receipt_collector_rows_from_fdb(date_from, date_to, user, password, collector):
    collector = normalize_collector_name(collector)
    if not collector:
        raise RuntimeError("Report 34 requires a collector. Example: python run_collection_query.py 34 2026-01-01 2026-01-31 angelique")

    sql = """
        SELECT
            p.PAYMENTDATE,
            COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), NULLIF(TRIM(p.USERID), ''), 'UNSPECIFIED') AS COLLECTOR_NAME,
            COALESCE(NULLIF(TRIM(p.AFTYPE), ''), NULLIF(TRIM(p.PAYGROUP_CT), ''), 'UNSPECIFIED') AS RECEIPT_TYPE,
            p.RECEIPTNO,
            COALESCE(NULLIF(TRIM(p.PAIDBY), ''), '-') AS TAXPAYER_NAME,
            p.STATUS_CT,
            p.VOID_BV,
            CASE
                WHEN COALESCE(p.PAYGROUP_CT, '') = 'RPT'
                    THEN COALESCE(rpt_totals.RPT_TOTAL, detail_totals.DETAIL_TOTAL, p.AMOUNT, 0)
                ELSE COALESCE(detail_totals.DETAIL_TOTAL, rpt_totals.RPT_TOTAL, p.AMOUNT, 0)
            END AS TOTAL_AMOUNT
        FROM PAYMENT p
        LEFT JOIN (
            SELECT PAYMENT_ID, SUM(AMOUNTPAID) AS DETAIL_TOTAL
            FROM PAYMENTDETAIL
            GROUP BY PAYMENT_ID
        ) detail_totals ON detail_totals.PAYMENT_ID = p.PAYMENT_ID
        LEFT JOIN (
            SELECT PAYMENT_ID, SUM(AMOUNT) AS RPT_TOTAL
            FROM PAYMENTCLASSDETAIL
            GROUP BY PAYMENT_ID
        ) rpt_totals ON rpt_totals.PAYMENT_ID = p.PAYMENT_ID
        WHERE p.PAYMENTDATE >= CAST(? AS DATE)
          AND p.PAYMENTDATE < DATEADD(1 DAY TO CAST(? AS DATE))
          AND UPPER(COALESCE(NULLIF(TRIM(p.COLLECTOR), ''), NULLIF(TRIM(p.USERID), ''), 'UNSPECIFIED')) = ?
        ORDER BY p.PAYMENTDATE, p.RECEIPTNO, p.PAYMENT_ID
    """
    rows = [["DATE", "Collector", "Receipt Type", "Receipt No", "Taxpayer name", "Status", "Total"]]
    con = connect_report_db(user, password, charset="WIN1252")
    try:
        cur = con.cursor()
        cur.execute(sql, (date_from, date_to, collector.upper()))
        for payment_date, collector_name, receipt_type, receipt_no, taxpayer_name, status_ct, void_bv, total_amount in cur.fetchall():
            rows.append([
                payment_date,
                collector_name,
                receipt_type,
                receipt_no,
                taxpayer_name,
                payment_status_label(status_ct, void_bv),
                total_amount or Decimal("0"),
            ])
        con.rollback()
    finally:
        con.close()
    return rows


def write_generate_receipt_collector_workbook(rows, output_path, date_from, date_to, collector):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Generate Receipt Collector"

    collector_label = normalize_collector_name(collector)
    body_rows = rows[1:]
    total_amount = Decimal("0")
    for row in body_rows:
        amount = row[6] or Decimal("0")
        if isinstance(amount, Decimal):
            total_amount += amount
        else:
            total_amount += Decimal(str(amount))

    thin = Side(style="thin", color="C9D5E3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EAF0F7")
    title_fill = PatternFill("solid", fgColor="123A5C")
    total_fill = PatternFill("solid", fgColor="EEF6F4")

    sheet.merge_cells("A1:G1")
    sheet["A1"] = "GENERATE RECEIPT COLLECTOR"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    sheet["A1"].fill = title_fill
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.merge_cells("A2:G2")
    sheet["A2"] = f"Collector: {collector_label}    Period: {date_from} to {date_to}    Includes paid, cancelled, and void payments"
    sheet["A2"].font = Font(italic=True, color="44546A")
    sheet["A2"].alignment = Alignment(horizontal="center")

    header_row = 4
    for col_index, value in enumerate(rows[0], start=1):
        cell = sheet.cell(header_row, col_index)
        cell.value = value
        cell.font = Font(bold=True, color="1F2D3D")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row_index, row_values in enumerate(body_rows, start=header_row + 1):
        for col_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row_index, col_index)
            cell.value = excel_value(value)
            cell.border = border
            cell.alignment = Alignment(vertical="top")
            if col_index == 1 and hasattr(value, "strftime"):
                cell.number_format = "yyyy-mm-dd"
            if col_index == 7:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")

    total_row = header_row + len(body_rows) + 1
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
    sheet.cell(total_row, 1).value = "TOTAL"
    sheet.cell(total_row, 1).font = Font(bold=True)
    sheet.cell(total_row, 1).fill = total_fill
    sheet.cell(total_row, 1).border = border
    sheet.cell(total_row, 1).alignment = Alignment(horizontal="right")
    total_cell = sheet.cell(total_row, 7)
    total_cell.value = excel_value(total_amount)
    total_cell.font = Font(bold=True)
    total_cell.fill = total_fill
    total_cell.border = border
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = Alignment(horizontal="right")

    for row in sheet.iter_rows(min_row=5, max_row=total_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border

    widths = {
        "A": 14,
        "B": 18,
        "C": 16,
        "D": 18,
        "E": 42,
        "F": 14,
        "G": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:G{max(total_row - 1, header_row)}"

    output_path = output_path.with_suffix(".xlsx")
    output_path = save_workbook_with_fallback(workbook, output_path)
    return len(body_rows), output_path


def run_fdb_template_report(number, output_path, date_from, date_to, user, password, collector=None):
    if number == 21:
        rows = build_full_summary_rows_from_fdb(date_from, date_to, user, password)
    elif number == 22:
        rows = build_no_rpt_rows_from_fdb(date_from, date_to, user, password)
    elif number == 23:
        rows = build_rpt_rows_from_fdb(date_from, date_to, user, password)
    elif number == 24:
        rows = build_rpt_detail_summary_from_fdb(date_from, date_to, user, password)
    elif number == 25:
        rows = build_rpt_record_rows_from_fdb(date_from, date_to, user, password)
    elif number == 26:
        rows = build_advance_rpt_record_rows_from_fdb(date_from, date_to, user, password)
    elif number == 27:
        rows = build_summary_sharing_rows_from_fdb(date_from, date_to, user, password)
    elif number == 28:
        rows = build_provincial_rpt_coding_rows_from_fdb(date_from, date_to, user, password)
    elif number == 29:
        rows, daily_rows = build_abstract_general_collections_rows_from_fdb(
            date_from, date_to, user, password
        )
    elif number == 30:
        rows, daily_rows = build_abstract_trust_funds_rows_from_fdb(
            date_from, date_to, user, password
        )
    elif number == 31:
        rows = build_full_report_collections_rows_from_fdb(date_from, date_to, user, password)
    elif number == 32:
        rows = build_cmci_annex_rows(date_from, date_to)
    elif number == 33:
        rows = build_tax_on_business_report(date_from, date_to)
    elif number == 34:
        rows = build_generate_receipt_collector_rows_from_fdb(date_from, date_to, user, password, collector)
    else:
        raise RuntimeError(f"Unsupported FDB template report {number}.")
    if number in (21, 22, 23):
        return write_template_workbook(number, rows, output_path, date_from)
    if number == 25:
        return write_rpt_record_workbook(rows, output_path, date_from, date_to)
    if number == 26:
        return write_advance_rpt_record_workbook(rows, output_path, date_from, date_to)
    if number == 27:
        return write_summary_sharing_workbook(rows, output_path, date_from, date_to)
    if number == 28:
        return write_provincial_rpt_coding_workbook(rows, output_path, date_from, date_to)
    if number == 29:
        return write_abstract_general_collections_workbook(
            rows, daily_rows, output_path, date_from, date_to
        )
    if number == 30:
        return write_abstract_trust_funds_workbook(
            rows, daily_rows, output_path, date_from, date_to
        )
    if number == 31:
        return write_full_report_collections_workbook(rows, output_path, date_from, date_to)
    if number == 32:
        return write_cmci_annex_workbook(rows, output_path, date_from, date_to)
    if number == 33:
        return write_tax_on_business_workbook(rows, output_path, date_from, date_to)
    if number == 34:
        return write_generate_receipt_collector_workbook(rows, output_path, date_from, date_to, collector)
    return write_csv_rows(output_path, rows)


def main():
    global CONNECTION_MODE
    parser = argparse.ArgumentParser(
        description="Run SELECT-only Firebird collection analysis queries and export CSV."
    )
    parser.add_argument(
        "query_number",
        nargs="?",
        type=int,
        help="Query number from collection_analysis_queries.sql. Use --list to see choices.",
    )
    parser.add_argument("date_from", nargs="?", type=parse_date, help="Start date, YYYY-MM-DD.")
    parser.add_argument("date_to", nargs="?", type=parse_date, help="End date, YYYY-MM-DD.")
    parser.add_argument("collector_name", nargs="?", help="Optional collector name for collector reports.")
    parser.add_argument("--list", action="store_true", help="List available queries and exit.")
    parser.add_argument(
        "--test-connection",
        action="store_true",
        help="Test the Firebird database connection and exit.",
    )
    parser.add_argument(
        "--connection",
        choices=("native", "odbc", "auto"),
        default=CONNECTION_MODE,
        help="Database connection mode. Use odbc for DSN=itaxzamboanguita.",
    )
    parser.add_argument("--collector", help="Optional collector/user id filter for collector reports.")
    parser.add_argument("--user", default="SYSDBA", help="Firebird username. Default: SYSDBA.")
    parser.add_argument("--password", default="masterkey", help="Firebird password. Default: masterkey.")
    args = parser.parse_args()
    CONNECTION_MODE = args.connection
    if args.collector is None and args.collector_name:
        args.collector = args.collector_name

    queries = load_queries()

    if args.list:
        list_queries(queries)
        return

    if args.test_connection:
        connection = open_database_connection(user=args.user, password=args.password, charset="UTF8")
        connection.close()
        print("STATUS: Connection test finished.")
        return

    if args.query_number is None or args.date_from is None or args.date_to is None:
        parser.error("query_number, date_from, and date_to are required unless --list is used.")

    selected = next((query for query in queries if query["number"] == args.query_number), None)
    selected_fdb_template_title = FDB_TEMPLATE_REPORTS.get(args.query_number)
    if selected is None and selected_fdb_template_title is None:
        parser.error(f"Query {args.query_number} was not found. Use --list to see choices.")

    if args.query_number == 34:
        args.collector = resolve_collector_selection(
            args.collector,
            args.date_from,
            args.date_to,
            args.user,
            args.password,
        )

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    collector_suffix = ""
    if args.collector:
        collector_suffix = "_" + re.sub(r"[^A-Za-z0-9_-]+", "_", args.collector.strip())
    output_path = OUTPUT_DIR / (
        f"query_{args.query_number:02d}{collector_suffix}_{args.date_from}_to_{args.date_to}.csv"
    )

    if selected_fdb_template_title is not None:
        row_count, output_path = run_fdb_template_report(
            args.query_number,
            output_path,
            args.date_from,
            args.date_to,
            args.user,
            args.password,
            args.collector,
        )
        print(f"Query {args.query_number}: {selected_fdb_template_title}")
        print(f"Rows exported: {row_count}")
        print(f"Output file: {output_path}")
        return

    sql = apply_parameters(selected["sql"], args.date_from, args.date_to, args.collector)

    statement_start = sql.lstrip().upper()
    if not (statement_start.startswith("SELECT") or statement_start.startswith("WITH")):
        raise RuntimeError("Refusing to run a non-SELECT statement.")

    connection = open_firebird_connection(
        user=args.user,
        password=args.password,
        charset="UTF8",
    )

    try:
        cursor = connection.cursor()
        cursor.execute(sql)
        row_count, output_path = export_rows(cursor, output_path)
        connection.rollback()
    finally:
        connection.close()

    print(f"Query {selected['number']}: {selected['title']}")
    print(f"Rows exported: {row_count}")
    print(f"Output file: {output_path}")


if __name__ == "__main__":
    main()

